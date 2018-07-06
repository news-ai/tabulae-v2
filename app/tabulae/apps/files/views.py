# -*- coding: utf-8 -*-
# Core Django imports
from django.contrib.auth import get_user_model
from django.shortcuts import get_object_or_404

# Third-party app imports
from openpyxl import load_workbook
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.decorators import list_route, detail_route
from rest_framework.filters import OrderingFilter

# Imports from app
from tabulae.apps.general.viewset import NewsAIModelViewSet
from tabulae.apps.general.response import Response, BulkResponse
from tabulae.apps.general.permissions import IsAdminOrIsSelf
from tabulae.apps.users.models import UserProfile
from tabulae.apps.contacts.models import Contact, CustomContactField
from tabulae.apps.lists.models import MediaList, CustomFieldsMap
from tabulae.apps.lists.serializers import MediaListSerializer
from tabulae.apps.publications.models import Publication
from .models import File
from .serializers import FileSerializer
from .permissions import FilePermission


class FileViewSet(NewsAIModelViewSet):
    serializer_class = FileSerializer
    permission_classes = (FilePermission,)
    filter_backends = (DjangoFilterBackend, OrderingFilter,)
    ordering_fields = ('created',)

    def _get_number_of_columns_for_file(self, ws_rows):
        '''
            File parsing method: Gets number of columns
            in a particular Excel worksheet.
        '''
        # Check where the starting position should be
        number_of_columns = len(ws_rows[0])
        starting_position = 0

        # If the first row is somehow corrupted
        # then we can just skip past it.
        if number_of_columns == 0:
            for i, row in enumerate(ws_rows):
                if len(row) > 0:
                    starting_position = i
                    number_of_columns = len(row)

        return starting_position, number_of_columns

    def _get_headers_with_file(self, ws):
        '''
            File parsing method: Returns the headers for a given
            Excel worksheet.
        '''
        headers = []
        ws_rows = list(ws.rows)
        if len(ws_rows) > 0:
            # Maximum number of rows we want to consider
            max_rows = 15
            if len(ws_rows) < max_rows + 1:
                max_rows = len(ws_rows)

            # Check where the starting position should be
            starting_position, number_of_columns = (
                self._get_number_of_columns_for_file(ws_rows))

            # Initialize a set of arrays
            for x in xrange(0, number_of_columns):
                headers.append([])
                headers[x] = {'rows': []}

            # Go through the sample rows
            for row in ws_rows[starting_position:max_rows]:
                single_row = []
                for i, column in enumerate(row):
                    # Set column value to be '' if it is
                    # not present
                    column_value = column.value
                    if column_value is None:
                        column_value = ''
                    else:
                        column_value = column_value.strip()

                    # Add as new column in that row
                    headers[i]['rows'].append(column_value)

        return headers

    def _get_headers(self, request, pk=None):
        '''
            File parsing method: Returns the headers for any
            given Excel file
        '''
        file = self.get_file_by_pk(request, pk)
        wb = load_workbook(file.file, read_only=True)

        headers = []
        if wb:
            # Get the active sheet (if any deleted)
            ws = wb.active

            # Get all headers by passing it the active worksheet
            headers = self._get_headers_with_file(ws)

        return headers

    def _is_custom_field(self, header_name):
        is_custom_field = True

        if header_name == 'firstname':
            is_custom_field = False
        elif header_name == 'lastname':
            is_custom_field = False
        elif header_name == 'email':
            is_custom_field = False
        elif header_name == 'notes':
            is_custom_field = False
        elif header_name == 'linkedin':
            is_custom_field = False
        elif header_name == 'twitter':
            is_custom_field = False
        elif header_name == 'instagram':
            is_custom_field = False
        elif header_name == 'website':
            is_custom_field = False
        elif header_name == 'blog':
            is_custom_field = False
        elif header_name == 'location':
            is_custom_field = False
        elif header_name == 'phonenumber':
            is_custom_field = False
        elif (header_name == 'employers' or
              header_name == 'pastemployers'):
            is_custom_field = False

        return is_custom_field

    def _contact_row_to_contact(self, request, file, contact_row):
        contact = Contact()
        custom_fields = []
        employers = []
        past_employers = []

        for i, row in enumerate(contact_row):
            value = row.value

            if value is None:
                value = ''
            else:
                value = value.strip()

            if file.order[i] != 'ignore_column':
                if file.order[i] == 'firstname':
                    contact.first_name = value
                elif file.order[i] == 'lastname':
                    contact.last_name = value
                elif file.order[i] == 'email':
                    contact.email = value
                elif file.order[i] == 'notes':
                    contact.notes = value
                elif file.order[i] == 'linkedin':
                    contact.linkedin = value
                elif file.order[i] == 'twitter':
                    contact.twitter = value
                elif file.order[i] == 'instagram':
                    contact.instagram = value
                elif file.order[i] == 'website':
                    contact.website = value
                elif file.order[i] == 'blog':
                    contact.blog = value
                elif file.order[i] == 'location':
                    contact.location = value
                elif file.order[i] == 'phonenumber':
                    contact.phone_number = value
                elif (file.order[i] == 'employers' or
                      file.order[i] == 'pastemployers'):
                    if value != '':
                        employer, created = Publication.objects.get_or_create(
                            name=value)

                        if file.order[i] == 'employers':
                            employers.append(employer)
                        else:
                            past_employers.append(employer)

                        if created:
                            employer.created_by = request.user
                            employer.save()
                else:
                    custom_field = CustomContactField()
                    custom_field.name = file.order[i]
                    custom_field.value = value
                    custom_field.created_by = request.user
                    custom_fields.append(custom_field)

        if len(custom_fields) > 0:
            custom_fields = CustomContactField.objects.bulk_create(
                custom_fields)

        contact.created_by = request.user
        user_profile = UserProfile.objects.get(user=request.user)
        contact.team = user_profile.team

        return (contact, custom_fields, employers, past_employers)

    def get_file_by_pk(self, request, pk):
        # Switch to team__pk=request.user.team.pk
        queryset = File.objects.filter(created_by=request.user)
        file = get_object_or_404(queryset, pk=pk)
        return file

    def retrieve(self, request, pk=None):
        if request.user and request.user.is_authenticated():
            file = self.get_file_by_pk(request, pk)
            serializer = FileSerializer(file)
            return Response(serializer.data, {})
        raise NotAuthenticated()

    def list(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return BulkResponse(serializer.data, {},
                            len(serializer.data), len(serializer.data))

    def get_queryset(self,):
        if self.request.user and self.request.user.is_authenticated():
            return File.objects.filter(
                created_by=self.request.user).order_by('-created')
        raise NotAuthenticated()

    def get_serializer(self, *args, **kwargs):
        if 'data' in kwargs:
            data = kwargs['data']

            if isinstance(data, list):
                kwargs['many'] = True

        return super(FileViewSet, self).get_serializer(*args, **kwargs)

    # GET /files/<id>/headers (External)
    @detail_route(methods=['get', 'post'], url_path='headers',
                  permission_classes=[IsAdminOrIsSelf])
    def headers(self, request, pk=None):
        if request.method == "GET":
            rows = self._get_headers(request, pk)
            return Response(rows, {})
        else:
            file = self.get_file_by_pk(request, pk)
            media_list = MediaList.objects.get(file=file)
            if ('headernames' in request.data and 'order' in request.data):
                # Save the file with the new information
                file.header_names = request.data['headernames']
                file.order = request.data['order']
                file.save()

                # Load workbook
                wb = load_workbook(file.file, read_only=True)
                if wb:
                    ws = wb.active

                    # Get starting position, and number of columns
                    ws_rows = list(ws.rows)
                    starting_position, number_of_columns = (
                        self._get_number_of_columns_for_file(ws_rows))

                    if len(file.header_names) != number_of_columns:
                        raise ParseError("Number of headers does "
                                         "not match the ones "
                                         "for the sheet")

                    # Parse the contacts
                    contacts = []
                    contacts_resources = []
                    for row in ws_rows[starting_position:]:
                        contact, custom_fields, employers, past_employers = (
                            self._contact_row_to_contact(
                                request, file, row))
                        if contact:
                            contacts.append(contact)
                            contacts_resources.append(
                                (custom_fields, employers, past_employers))

                    if len(contacts) > 0:
                        if len(contacts) > 0:
                            new_contacts = Contact.objects.bulk_create(
                                contacts)

                            for i, contact in enumerate(new_contacts):
                                if contacts_resources[i]:
                                    # index 0: custom_fields
                                    if len(contacts_resources[i][0]) > 0:
                                        contact.custom_fields.set(
                                            contacts_resources[i][0])

                                    # index 1: employers
                                    if len(contacts_resources[i][1]) > 0:
                                        contact.employers.set(
                                            contacts_resources[i][1])

                                    # index 2: past_employers
                                    if len(contacts_resources[i][2]) > 0:
                                        contact.past_employers.set(
                                            contacts_resources[i][2])

                        for i, header in enumerate(file.order):
                            if header != 'ignore_column':
                                is_custom_field = self._is_custom_field(header)
                                if is_custom_field:
                                    custom_field = CustomFieldsMap()
                                    custom_field.name = file.header_names[i]
                                    custom_field.value = header
                                    custom_field.custom_field = True
                                    custom_field.hidden = False
                                    custom_field.save()

                                    media_list.fields_map.add(custom_field)

                        media_list.contacts.set(new_contacts)
                        media_list.save()

                serializer = MediaListSerializer(media_list)
                return Response(serializer.data, {})
            raise ParseError()

    # GET /files/<id>/sheets (External)
    @detail_route(methods=['get'], url_path='sheets',
                  permission_classes=[IsAdminOrIsSelf])
    def sheets(self, request, pk=None):
        sheet_names = []
        file = self.get_file_by_pk(request, pk)
        wb = load_workbook(file.file, read_only=True)
        if wb:
            sheetnames = wb.get_sheet_names()

        data = {
            'names': sheetnames
        }
        return Response(sheetnames, {})
